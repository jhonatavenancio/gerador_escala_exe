import tkinter as tk
from tkinter import messagebox
from pandas.tseries.offsets import MonthEnd
import pandas as pd
import random
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Função para gerar uma lista de datas em um mês específico
def generate_workdays(year, month):
    start_date = datetime(year, month, 1)
    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)
    workdays = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
    workdays = [day for day in workdays if day.weekday() < 5]  # Excluir fins de semana
    return workdays

# Função para definir as semanas
def define_weeks(workdays):
    weeks = {}
    current_week = []
    last_weekday = -1
    week_number = 0

    for workday in workdays:
        weekday = workday.weekday()
        if weekday <= last_weekday:
            weeks[week_number] = current_week
            current_week = []
            week_number += 1
        current_week.append(workday)
        last_weekday = weekday
    weeks[week_number] = current_week
    return weeks

# Função para gerar a escala para um mês específico
def generate_schedule(employee_names, year, month):
    workdays = generate_workdays(year, month)
    weeks = define_weeks(workdays)

    df = pd.DataFrame(index=workdays, columns=employee_names)
    df.index.name = 'Dia'
    df.reset_index(inplace=True)
    df['Dia'] = df['Dia'].dt.strftime("%d/%m/%Y")

    # Congele a primeira linha
    df.freeze_panes = 'A2'

    # Passo 1: Dar a todos 3 dias de home office e 2 presenciais por semana
    for employee in employee_names:
        for week in weeks.values():
            if len(week) >= 5:
                home_office_days = random.sample(week, 3)
                presencial_days = [day for day in week if day not in home_office_days]
            else:
                home_office_days = random.sample(week, 1)
                presencial_days = week  # Para semanas curtas

            for day in week:
                df.loc[df['Dia'] == day.strftime("%d/%m/%Y"), employee] = (
                    "Home" if day in home_office_days else "Presencial"
                )

    # Passo 2: Garantir pelo menos um funcionário presencial todos os dias
    for index, row in df.iterrows():
        if "Presencial" not in row.values:
            df.loc[index, random.choice(employee_names)] = "Presencial"

    # Passo 3: Evitar que o mesmo resultado ocorra mais de duas vezes seguidas para um funcionário
    for employee in employee_names:
        streak = {"Home": 0, "Presencial": 0}
        for index, row in df.iterrows():
            if row[employee] == "Home":
                streak["Home"] += 1
                streak["Presencial"] = 0
            else:
                streak["Presencial"] += 1
                streak["Home"] = 0

            # Se houver mais de 2 dias presenciais seguidos, alterne
            if streak["Presencial"] > 2:
                df.loc[index, employee] = "Home"
                streak["Home"] = 1
                streak["Presencial"] = 0

            # Se houver mais de 2 dias de home office seguidos, alterne
            if streak["Home"] > 2:
                df.loc[index, employee] = "Presencial"
                streak["Home"] = 0
                streak["Presencial"] = 1

    return df, weeks

# Função para salvar a escala em arquivo Excel
def save_schedule(schedule, weeks):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    target_folder = os.path.expanduser("~/Documentos/home_presencial")
    if not os.path.exists(target_folder):
        os.makedirs(target_folder)
    file_path = os.path.join(target_folder, f"escala_trabalho_{timestamp}.xlsx")
    schedule.to_excel(file_path, index=False)

    # Adicionar cores de semana ao Excel
    week_colors = [
        PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type="solid"),  # Azul para semana 1
        PatternFill(start_color='FA8072', end_color='FA8072', fill_type="solid"),  # Vermelho para semana 2
        PatternFill(start_color='90EE90', end_color='90EE90', fill_type="solid"),  # Verde para semana 3
        PatternFill(start_color='FFD700', end_color='FFD700', fill_type="solid"),  # Amarelo para semana 4
        PatternFill(start_color='9370DB', end_color='9370DB', fill_type="solid"),  # Roxo para semana 5
    ]

    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    workbook = load_workbook(file_path)
    worksheet = workbook.active
    for week_index, week in weeks.items():
        if week_index < len(week_colors):
            color = week_colors[week_index]
            for workday in week:
                # Procura a linha correspondente ao dia
                for row in worksheet.iter_rows():
                    if row[0].value == workday.strftime("%d/%m/%Y"):
                        for cell in row:
                            cell.fill = color
                            cell.border = thin_border  # Adicionar bordas finas
                        break

    workbook.save(file_path)
    return file_path

# Função para gerar a escala e exibir uma mensagem de sucesso ou erro
def generate_and_save_schedule(employee_names_str, year, month):
    employee_names = [name.strip() for name in employee_names_str.split(",")]
    if len(employee_names) > 20:
        raise ValueError("Por favor, insira no máximo 20 nomes.")

    schedule, weeks = generate_schedule(employee_names, year, month)

    return save_schedule(schedule, weeks)

# Interface gráfica usando tkinter
class ScheduleApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gerador de Escala de Trabalho")

        self.label_employee_names = tk.Label(self, text="Nomes dos funcionários (separados por vírgula):")
        self.label_employee_names.grid(row=0, column=0, padx=10, pady=5)

        self.entry_employee_names = tk.Entry(self, width=50)
        self.entry_employee_names.grid(row=0, column=1, padx=10, pady=5)

        self.label_month = tk.Label(self, text="Mês (Abril = 4):")
        self.label_month.grid(row=1, column=0, padx=10, pady=5)

        self.entry_month = tk.Entry(self, width=11)
        self.entry_month.grid(row=1, column=1, padx=10, pady=5)

        self.label_year = tk.Label(self, text="Ano (2024):")
        self.label_year.grid(row=2, column=0, padx=10, pady=5)

        self.entry_year = tk.Entry(self, width=11)
        self.entry_year.grid(row=2, column=1, padx=10, pady=5)

        self.button_generate = tk.Button(self, text="Gerar Escala", command=self.generate_schedule)
        self.button_generate.grid(row=3, column=0, columnspan=2, pady=10)

                # Botão para exibir as instruções
        self.button_instructions = tk.Button(self, text="Instruções", command=self.show_instructions)
        self.button_instructions.grid(row=4, column=0, columnspan=2, pady=10)

    def show_instructions(self):
        # Crie uma nova janela para exibir as instruções
        instructions_window = tk.Toplevel(self)
        instructions_window.title("Instruções")

        # Adicione um rótulo com as informações do gerador de planilhas
        instructions_label = tk.Label(instructions_window, text="Instruções do Gerador de Planilhas:\n\n"
                                                                "- Todos os funcionários devem ter 3 dias de HomeOffice e 2 dias presenciais.\n"
                                                                "- A semana deve ser separada por cores.\n"
                                                                "- O arquivo deve ser salvo em: Documentos/home_presencial\n\n"
                                                                "Desenvolvido por: Jhonata Venâncio")
        instructions_label.pack(padx=10, pady=10)

    def generate_schedule(self):
        try:
            employee_names_str = self.entry_employee_names.get()
            month = int(self.entry_month.get())
            year = int(self.entry_year.get())
            file_path = generate_and_save_schedule(employee_names_str, year, month)
            messagebox.showinfo("Sucesso", f"Escala de trabalho salva em '{file_path}'")
        except Exception as e:
            messagebox.showerror("Erro", str(e))



if __name__ == "__main__":
    app = ScheduleApp()
    app.mainloop()
